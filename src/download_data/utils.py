import datetime
import os

from dls.settings import MEDIA_ROOT

from openpyxl import Workbook
from openpyxl.styles import (Alignment,
                             Border,
                             Font,
                             NamedStyle,
                             PatternFill,
                             Side)


class ExcelFileMaker:
    data: list
    columns: list
    filename: str
    filepath: str
    filepath_db: str

    def __init__(self, data: list = None, columns: list = None,
                 filename: str = None):
        self.data = data if data is not None else []
        self.set_columns(columns)
        self.set_filename(filename)
        self.make_file()

    def set_filename(self, filename: str = None):
        if filename is not None:
            self.filename = f'{filename}.xlsx'
        else:
            self.filename = (f'generated_'
                             f'{datetime.date.today().strftime("%d.%m.%Y")}'
                             f'.xlsx')
        self.filepath = os.path.join(MEDIA_ROOT, "downloaded_data",
                                     self.filename)
        self.filepath_db = os.path.join("downloaded_data", self.filename)
        counter = 0
        while os.path.isfile(self.filepath):
            counter += 1
            if filename is not None:
                self.filename = f'{filename}({counter}).xlsx'
            else:
                self.filename = (
                    f'generated_'
                    f'{datetime.date.today().strftime("%d.%m.%Y")}'
                    f'({counter}).xlsx'
                )
            self.filepath = os.path.join(MEDIA_ROOT, "downloaded_data",
                                         self.filename)
            self.filepath_db = os.path.join("downloaded_data", self.filename)

    def set_columns(self, columns: list = None):
        if columns is not None:
            self.columns = columns
        else:
            self.columns = []
            for key in self.data:
                self.columns.append(key)

    def make_file(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Книга 1"
        header = self.columns
        ws.append(header)
        for row in self.data:
            ws.append([row[col] for col in header])

        header_style = NamedStyle(name='header')
        header_style.font = Font(bold=True, color='FFFFFF')
        header_style.alignment = Alignment(horizontal='center',
                                           vertical='center')
        header_style.fill = PatternFill(start_color='4F81BD',
                                        end_color='4F81BD', fill_type='solid')
        border_style = Border(
            left=Side(border_style='thin',
                      color='000000'),
            right=Side(border_style='thin',
                       color='000000'),
            top=Side(border_style='thin',
                     color='000000'),
            bottom=Side(border_style='thin',
                        color='000000')
        )
        header_style.border = border_style

        cell_style = NamedStyle(name='cell')
        cell_style.alignment = Alignment(horizontal='left',
                                         vertical='center')
        cell_style.border = border_style

        for cell in ws[1]:
            cell.style = header_style

        for row in ws.iter_rows(min_row=2,
                                max_row=ws.max_row,
                                min_col=1,
                                max_col=ws.max_column):
            for cell in row:
                cell.style = cell_style

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(self.filepath)
